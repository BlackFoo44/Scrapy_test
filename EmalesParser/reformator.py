import csv

import openpyxl

wb = openpyxl.Workbook()
ws = wb.active
ws1 = wb.create_sheet("16_elem", 0)
ws2 = wb.create_sheet("all_enamels", 1)
ws3 = wb.create_sheet("all_items", 2)

try:
    with open("first_16_enamels.csv", encoding="utf8") as f:
        reader = csv.reader(f, delimiter=",")
        for row in reader:
            ws1.append(row)
except FileNotFoundError:
    print("Файл first_16_element не найден")
try:
    with open("all_enamels.csv", encoding="utf8") as f:
        reader = csv.reader(f, delimiter=",")
        for row in reader:
            ws2.append(row)
except FileNotFoundError:
    print("Файл all_enamels не найден")
try:
    with open("all_items.csv", encoding="utf8") as f:
        reader = csv.reader(f, delimiter=",")
        for row in reader:
            ws3.append(row)
    wb.save("parser.xlsx")
except FileNotFoundError:
    print("Файл all_items не найден")